"""Lectura de archivos tabulares (plan de documentos, E3).

CSV con la stdlib y XLSX con openpyxl → una matriz de celdas (list[list[str]]),
sin interpretar nada: la normalización (qué columna es qué, formato de fecha,
signo) vive en la capa de aplicación (estado_cuenta.py). Aquí solo se lee.

openpyxl es Python puro (sin binarios nativos) → seguro para la imagen de
Railway (riesgo R8: nada de camelot/opencv).
"""

from __future__ import annotations

import csv
import io
import logging

log = logging.getLogger("e5.tabular")


def leer_tabla(contenido: bytes, content_type: str, filename: str | None = None) -> list[list[str]]:
    """Archivo tabular → filas de celdas como texto. Filas vacías descartadas.
    Ante cualquier fallo devuelve [] (el llamador lo trata como respaldo)."""
    try:
        if _es_xlsx(content_type, filename):
            return _leer_xlsx(contenido)
        return _leer_csv(contenido)
    except Exception:
        log.warning("No se pudo leer el archivo tabular (%s).", content_type, exc_info=True)
        return []


def _es_xlsx(content_type: str, filename: str | None) -> bool:
    if "spreadsheetml" in content_type or "ms-excel" in content_type:
        return True
    return bool(filename and filename.lower().endswith((".xlsx", ".xls")))


def _leer_xlsx(contenido: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    hoja = wb.active
    filas: list[list[str]] = []
    for fila in hoja.iter_rows(values_only=True):
        celdas = ["" if c is None else str(c).strip() for c in fila]
        if any(celdas):
            filas.append(celdas)
    wb.close()
    return filas


def _leer_csv(contenido: bytes) -> list[list[str]]:
    texto = _decodificar(contenido)
    # Sniff del delimitador (Ecuador: coma o punto y coma según el banco).
    muestra = texto[:2048]
    try:
        dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
        delimitador = dialecto.delimiter
    except csv.Error:
        delimitador = ";" if muestra.count(";") > muestra.count(",") else ","
    filas: list[list[str]] = []
    for fila in csv.reader(io.StringIO(texto), delimiter=delimitador):
        celdas = [c.strip() for c in fila]
        if any(celdas):
            filas.append(celdas)
    return filas


def _decodificar(contenido: bytes) -> str:
    """Los exports de banca ecuatoriana suelen venir en latin-1/cp1252; se
    intenta utf-8 primero y se cae a latin-1 (nunca falla)."""
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return contenido.decode(enc)
        except UnicodeDecodeError:
            continue
    return contenido.decode("latin-1", errors="replace")
